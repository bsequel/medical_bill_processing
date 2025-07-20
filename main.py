
from aws_lib_.aws_ocr_main import main_call
import os
import shutil
from openpyxl import Workbook
import csv
from openpyxl import load_workbook
import pandas as pd
from collections import defaultdict
import config
from db_operations import all_data ,  excel_data

conn = config.conn
cur = config.cur

input_folder = r"G:\SS\DFCCIL\Medical_project\input"
output_folder_path = r"G:\SS\DFCCIL\Medical_project\output"
excel_path = r"G:\SS\DFCCIL\Medical_project\excel"

print("Medical Boat is started.................")



def Trigger(input_path):
    output_path = output_folder_path    
    text=''
    os.chdir(output_path)
    main_call(input_path)
    text_all=''
    for file in os.listdir(output_path):
        if file.endswith('text.txt'):
            output_file = os.path.join(output_path,file)
            text_=open(output_file,'r',encoding="utf-8")
            text_=text_.readlines()
            text_=' '.join(text_)
            text_all = text_all + text_

    wb = Workbook()
    ws = wb.active
    for file in os.listdir(output_path): #output
        if file.endswith('tables.csv'):
            
            csv_file_path = os.path.join(output_path,file)
            with open(csv_file_path, 'r') as f:
                for row in csv.reader(f):
                    ws.append(row)
                    excel_file_path = excel_path+"\\"+"output.xlsx"
                    wb.save(excel_file_path)

    return text_all

# text = Trigger(input_folder)
excel_path = r"G:\SS\DFCCIL\Medical_project\master_excel\room.xlsx"
level_excel = r"G:\SS\DFCCIL\Medical_project\master_excel\level_data.xlsx"
level_excel_wb = load_workbook(level_excel)
level_excel_ws = level_excel_wb.active
import time 
time.sleep(10)
# print("Data Extracted successfully.................")



def level_compare():
    # level = find_level(text)
    # if level == 'Jr. Exe.':
    level = '/Executive'
    for i in range(1,level_excel_ws.max_row):
        if level_excel_ws[f'B{i}'].value is not None:
            if level in level_excel_ws[f'B{i}'].value:
                # print(level_excel_ws[f'B{i}'].value,'+++++++++++++++++++++++',level_excel_ws[f'C{i}'].value)
                room_type = level_excel_ws[f'C{i}'].value
    return room_type



def room_rent():
    # wb = load_workbook(filename=file_path, data_only=True)
    # ws = wb[sheet_name]
    
    result = []
    room_rent_count = 0
    capture = False
    empty_count = 0
    unique_data = {}
    count_dict = defaultdict(int)

    
    col_B, col_C, col_E, col_F, col_G = 2, 3, 5, 6, 7  # Column indices (1-based)
    
    for row in level_excel_ws.iter_rows(min_row=1, values_only=True):
        # print(row[col_B - 1])
        if row[col_B - 1] is not None or row[col_B - 1] == ' ':
            if row[col_B - 1].strip()== "ROOM RENT":
                print('###############')
                room_rent_count += 1
                empty_count = 0
                if room_rent_count == 2:
                    capture = True
                continue
            
        if capture:
            if row[col_B - 1] is None:
                empty_count += 1
              


            else:
                empty_count = 0
            
            if empty_count >= 2:
                break
            
            result.append((row[col_C - 1], row[col_E - 1], row[col_F - 1], row[col_G - 1]))
    # print(result)
    
    filtered_data = [row for row in result if not all(x is None for x in row)]
    for row in filtered_data:
        bed_type = row[0]  # First column value (Bed Type)
        if bed_type:  # Ensure column C has a valid value
            count_dict[bed_type] += 1
            if bed_type not in unique_data:
                unique_data[bed_type] = row
    
    return filtered_data , dict(count_dict)

shutil.copy(r"G:\SS\DFCCIL\Medical_project\aws_lib_\final_excel\final.xlsx", r"G:\SS\DFCCIL\Medical_project\master_excel/bill.xlsx")
time.sleep(10)
all_data(conn,cur)
import shutil
inputPath = r"G:\SS\DFCCIL\Medical_project\input"
for file in os.listdir(inputPath):
    filePath = os.path.join(inputPath, file)
    fileName = os.path.basename(filePath)
    if fileName == "202503180329531.pdf":
        shutil.copy(r"G:\SS\DFCCIL\Medical_project\aws_lib_\final.xlsx",r"G:\DFCCIL\Sheet\final.xlsx")
        # excel_data()
        # time.sleep(10)
        print("Sheet Created .................")
    else:
        # time.sleep(10)
        print("Missing data from Max Sheet .................")
        print("Sheet Not Created .................")

    
    














    






