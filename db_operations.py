import pandas as pd
import psycopg2

# Read the Excel file



# Connect to PostgreSQL

import pandas as pd
import psycopg2
from openpyxl.styles import Alignment

def all_data(conn, cur):
    try:
        # Load Excel files
        df = pd.read_excel(r'G:\SS\DFCCIL\Medical_project\master_excel\master_data.xlsx')
        room_rent = pd.read_excel(r'G:\SS\DFCCIL\Medical_project\master_excel\room_rent.xlsx')
        level = pd.read_excel(r'G:\SS\DFCCIL\Medical_project\master_excel\level_data.xlsx')

        batch_size = 1000  # Batch size for bulk inserts

        # Create Table: others_procedures
        cur.execute("""
        CREATE TABLE IF NOT EXISTS others_procedures (
            Department VARCHAR(256),
            Item_ID VARCHAR(100),
            Item_Name VARCHAR(256),
            OP VARCHAR(100),
            Daycare VARCHAR(100),
            Economy VARCHAR(100),
            Double VARCHAR(100),
            Single_ICU VARCHAR(100),
            Classic_Deluxe VARCHAR(100),
            Suite VARCHAR(100)
        );
        """)
        cur.execute("TRUNCATE TABLE others_procedures")
        conn.commit()

        # Insert data into others_procedures
        data_tuples = [tuple(x) for x in df.to_numpy()]
        insert_query = """
        INSERT INTO others_procedures (
            Department, Item_ID, Item_Name, OP, Daycare, Economy, Double, Single_ICU, Classic_Deluxe, Suite
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        for i in range(0, len(data_tuples), batch_size):
            batch = data_tuples[i:i + batch_size]
            cur.executemany(insert_query, batch)
        conn.commit()

        # Create Table: room_rent
        cur.execute("""
        CREATE TABLE IF NOT EXISTS room_rent (
            room_type VARCHAR(100),
            room_rent_per_day VARCHAR(100),
            Consultation_Charges_per_visit VARCHAR(100)
        );
        """)
        cur.execute("TRUNCATE TABLE room_rent")
        conn.commit()

        # Insert data into room_rent
        room_data = [tuple(x) for x in room_rent.to_numpy()]
        insert_query = """
        INSERT INTO room_rent (room_type, room_rent_per_day, Consultation_Charges_per_visit) 
        VALUES (%s, %s, %s);
        """
        for i in range(0, len(room_data), batch_size):
            batch = room_data[i:i + batch_size]
            cur.executemany(insert_query, batch)
        conn.commit()

        # Create Table: level_data
        cur.execute("""
        CREATE TABLE IF NOT EXISTS level_data (
            S_NO INT,
            levels VARCHAR(100),
            room_type VARCHAR(100)
        );
        """)
        cur.execute("TRUNCATE TABLE level_data")
        conn.commit()

        # Insert data into level_data
        level_data = [tuple(x) for x in level.to_numpy()]
        insert_query = """
        INSERT INTO level_data (S_NO, levels, room_type) VALUES (%s, %s, %s);
        """
        for i in range(0, len(level_data), batch_size):
            batch = level_data[i:i + batch_size]
            cur.executemany(insert_query, batch)
        conn.commit()

        print("Data successfully inserted into others_procedures, room_rent, and level_data.")

        # ================== BILL DATA PROCESSING ==================
        file_path = r"G:\SS\DFCCIL\Medical_project\master_excel\bill.xlsx"
        xls = pd.ExcelFile(file_path)
        df_bill = pd.read_excel(xls, sheet_name="Sheet1", skiprows=4, dtype=str)

        # Rename and clean up data
        df_bill.columns = ["Date", "Particulars", "Service_Code", "Rate", "Qty", "Amount"]
        df_bill = df_bill.iloc[1:].reset_index(drop=True)
        df_bill["Service"] = df_bill["Date"].where(df_bill["Particulars"].isna()).ffill()
        total_rows = df_bill["Particulars"].str.contains("Total", case=False, na=False)
        service_totals = {df_bill.loc[idx - 1, "Service"]: df_bill.loc[idx, "Amount"] for idx in df_bill[total_rows].index}
        df_bill["Total"] = df_bill["Service"].map(service_totals)
        df_cleaned = df_bill[~total_rows].dropna(subset=["Particulars"]).reset_index(drop=True)
        df_cleaned["Rate"] = pd.to_numeric(df_cleaned["Rate"], errors="coerce")
        df_cleaned["Qty"] = pd.to_numeric(df_cleaned["Qty"], errors="coerce")
        df_cleaned["Amount"] = pd.to_numeric(df_cleaned["Amount"], errors="coerce")

        # Create Table: services
        cur.execute("""
        CREATE TABLE IF NOT EXISTS services (
            id SERIAL PRIMARY KEY,
            date TEXT,
            particulars TEXT,
            service_code TEXT,
            rate VARCHAR(100),
            qty VARCHAR(100),
            amount VARCHAR(100),
            service TEXT,
            total VARCHAR(100)
        );
        """)
        cur.execute("TRUNCATE TABLE services")
        conn.commit()

        # Insert data into services
        insert_query = """
        INSERT INTO services (date, particulars, service_code, rate, qty, amount, service, total)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
        """
        service_data = [tuple(row) for row in df_cleaned.to_numpy()]
        for i in range(0, len(service_data), batch_size):
            batch = service_data[i:i + batch_size]
            cur.executemany(insert_query, batch)
        conn.commit()

        # print("Data successfully inserted into services.")

    except Exception as e:
        print("Error:", e)

    finally:
        cur.close()
        conn.close()







from openpyxl import Workbook
from openpyxl.styles import Alignment

def excel_data():
    wb = Workbook()
    ws = wb.active







    
    
    ws['A1']='PROCESSING SHEET FOR MEDICAL'
    ws.merge_cells("A1:H1")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells("A2:D2")
    ws.merge_cells("E2:H2")
    ws['A2']='Name & Designation of Officer:  '
    ws['E2']='Smt. Durga Bishwakarma, Jr. Executive'
    
    ws.merge_cells("A3:D3")
    ws.merge_cells("E3:H3")
    ws['A3']='Employee Code:        '
    ws['E3']='103682'
    
    ws.merge_cells("A4:D4")
    ws.merge_cells("E4:H4")
    ws['A4']='Name & Relation with the patient:'
            
    ws['E4']='Son'
    
    ws['A5']='Date of Admission'
    ws['E5']='23-11-2024'
    
    ws['A6']='Date of Discharge'
    ws['E6']='01-12-2024'

    ws.merge_cells("A5:D5")
    ws.merge_cells("A6:D6")
    ws.merge_cells("E5:H5")
    ws.merge_cells("E6:H6")


    
    
    columns = ["S.no","Bill No.", "Date",  "Breakup", "Claimed amount","Quantity", "MaxSaket Rates", "Total Amount Admissible", "Net Admissible"]
    # ws[f"C{8}"] = "Room Rent"
    # ws.merge_cells('C8:C16')  # ✅ Correct

    ws["C8"].alignment = Alignment(horizontal="center", vertical="center")
    specified_row = 7
    
    for col_num, value in enumerate(columns, start=1):  # Start from column 1 (A)
        ws.cell(row=specified_row, column=col_num, value=value)
    a=8
    for i in range (a,a+190):
        ws[f'A{i}']=i-7
        # ws[f'B{i}']='01-12-2024'
        ws[f'B{i}']='IPCA325/228704'
    
    data = [
        ["23/11/2024","Bed Charges PAED ICU", 10000,  1,  "9000/-", 9000, 9000],
        ["23/11/2024","Bed Charges PAED ICU", 10000, 1,  "9000/-", 9000, 9000],
        ["24/11/2024","Bed Charges PAED ICU", 10000, 1,  "9000/-", 9000, 9000],
        ["25/11/2024","Bed Charges PAED ICU", 10000, 1,  "9000/-", 9000, 9000],
        ["26/11/2024","Bed Charges General", 3000, 1,  4500, 3000, 3000],
        ["27/11/2024","Bed Charges General", 3000, 1,  4500, 3000, 3000],
        ["28/11/2024","Bed Charges General", 3000, 1,  4500, 3000, 3000],
        ["29/11/2024","Bed Charges General", 3000, 1,  4500, 3000, 3000],
        ["30/11/2024","Bed Charges General", 3000,  1, 4500, 3000, 3000],


        ["23/11/2024","CBC", 440, 1, 1100, 440, 440],  
        ["25/11/2024","CBC", 440, 1, 1100, 440, 440],
        ["29/11/2024","CBC", 440, 1, 1100, 440, 440],
        ["23/11/2024","CRP", 770, 1, 750, 770, 750],
        ["29/11/2024","CRP", 770, 1, 750, 770, 750],
        ["23/11/2024","UREA AND CREATININE", 990, 1, 1220, 990, 990],
        ["23/11/2024","LIVER PROFILE", 1760, 1, 2580, 1760, 1760],
        ["23/11/2024","SERUM SODIUM", 770, 1, 750, 750, 750],
        ["25/11/2024","SERUM SODIUM", 770, 1, 750, 750, 750],
        ["23/11/2024","POTASSIUM SODIUM", 660, 1, 700, 660, 660],
        ["25/11/2024","POTASSIUM SODIUM", 660, 1, 700, 660, 660],
        ["23/11/2024","CALCIUM", 660, 1, 680, 660, 660],
        




        ["23/11/2024","X-RAY: ABDOMEN", 550, 1, 950, 550, 550],
        ["25/11/2024","X-RAY: ABDOMEN", 550, 1, 950, 550, 550],


        ["23/11/2024","Consultation", 1000, 1, 1650, 1000, 1000],
        ["23/11/2024","Follow Consultation", 1000, 1, 1650, 1000, 1000],

        ["24/11/2024","Consultation", 1000, 1, 1650, 1000, 1000],
        ["24/11/2024","Follow Consultation", 1000, 1, 1650, 1000, 1000],

        ["25/11/2024","Consultation", 1000, 1, 1650, 1000, 1000],
        ["25/11/2024","Follow Consultation", 1000, 1, 1650, 1000, 1000],

        ["26/11/2024","Consultation", 500, 1, 1650, 500, 500],
        ["26/11/2024","Follow Consultation", 500, 1, 1650, 500, 500],
        ["27/11/2024","Consultation", 500, 1, 1650, 500, 500],
        ["27/11/2024","Follow Consultation", 500, 1, 1650, 500, 500],
        ["28/11/2024","Consultation", 500, 1, 1650, 500, 500],
        ["28/11/2024","Follow Consultation", 500, 1, 1650, 500, 500],
        ["29/11/2024","Consultation", 500, 1, 1650, 500, 500],
        ["29/11/2024","Follow Consultation", 500, 1, 1650, 500, 500],
        ["30/11/2024","Consultation", 500,1, 1650, 500, 500],
        ["30/11/2024","Follow Consultation", 500,1, 1650, 500, 500],

        ["01/12/2024","Consultation", 500,1, 1650, 500, 500],


        ["23/11/2024","DNS 500ML PLASTIC", 42.80, 2, '-', 85.60, 85.60],
        ["23/11/2024","EMESET 2ML INJ (2 MG/1 ML)", 13.35, 3, '-', 40.05, 40.05],
        ["23/11/2024","SODIUM CHLORIDE 0.9% 100ML (NIRLIFE)", 22.03, 3, '-', 65.09, 65.09],
        ["23/11/2024","MONOTAX 1GM/VIAL INJ ", 69.86, 2, '-', 139.72, 139.72],
        ["23/11/2024","RLOC 2ML INJ (25MG/ML)  ", 7.26, 2, '-', 14.52, 14.52],
        ["23/11/2024","MEROZA 500MG INJ   ", 809.77, 2, '-', 1619.54, 1619.54],
        ["23/11/2024","PARACETAMOL (BBRAUN) 100ML IV INJ    ", 654.50, 2, '-', 1309.00, 1309.00],
        ["23/11/2024","RLOC 2ML INJ (25MG/ML) ",7.26, 3, '-', 21.78, 21.78],
        ["23/11/2024","AMIKAMAC 100MG/2ML INJ VIAL ",34.21, 2, '-', 68.42, 68.42],
        ["23/11/2024","METRONIDAZOLE 100ML IV (NIRLIFE)_ ",23.52, 2, '-', 47.04, 47.04],
        ["23/11/2024","NUFORCE DUSTING POWDER 75GM ",104.50, 1, '-', 104.50, 104.50],
        ["24/11/2024","DNS 500ML PLASTIC (NIRLIFE)  ",42.80, 2, '-', 85.60, 85.60],
        ["24/11/2024","LASIX 40MG/ 4ML INJ (10 MG/1 ML). ",13.39, 1, '-', 13.39, 13.39],
        ["24/11/2024","MEROZA 500MG INJ  ",809.77, 1, '-', 809.77, 809.77],
        ["24/11/2024","SODIUM CHLORIDE 0.9% 100ML (NIRLIFE)",22.03, 2, '-', 44.06, 44.06],
        ["24/11/2024","VC-VLOK 4MG/2ML INJ",15.44, 1, '-',15.44,15.44],
        ["24/11/2024","RLOC 2ML INJ (25MG/ML) ",7.26, 1, '-',7.26,7.26],
        ["24/11/2024","AMIKAMAC 100MG/2ML INJ VIAL ",34.21, 1, '-',34.21,34.21],
        ["25/11/2024","DNS 500ML PLASTIC (NIRLIFE)",42.80, 1, '-',42.80,42.80],
        ["25/11/2024","EMESET 2ML INJ (2 MG/1 ML) ",13.35, 3, '-',40.05,40.05],
        ["25/11/2024","MEROZA 500MG INI ",809.77, 1, '-',809.77,809.77],
        ["25/11/2024","SODIUM CHLORIDE 0.9% 100ML (NIRLIFE)  ",22.03, 2, '-',44.06,44.06],
        ["25/11/2024","RLOC 2ML INJ (25MG/ML) ",7.26, 2, '-',14.52,14.52],
        ["26/11/2024","DNS 500MI PLASTIC (NIRLIFE) ",42.80, 1, '-',42.80,42.80],
        ["26/11/2024","MONOTAX 1GM/VIAL INJ ",69.86, 1, '-',69.86,69.86],
        ["26/11/2024","EMESET 2ML INJ (2 MG/1 ML) ",13.35, 2, '-',26.70,26.70],
        ["26/11/2024","MEROZA 500MG INJ ",809.77, 1, '-',809.77,809.77],
        ["26/11/2024","METRONIDAZOLE 100ML IV (NIRLIFE).",23.52, 1, '-',23.52,23.52],
        ["26/11/2024","PARACETAMOL (BBRAUN)100ML IV IN) ",654.50, 1, '-',654.50,654.50],
        ["26/11/2024","RLOC 2ML IN) (25MG/ML) ",7.26 , 2, '-',14.52,14.52],
        ["26/11/2024","AMIKAMAC 100MG/2ML INJ VIAL ",34.21 , 1, '-',34.21,34.21],
        ["26/11/2024","MAMY POKO PANTS-M6(01PKT-06PCS) ",99.00 , 2, '-',198.00,198.00],
        ["27/11/2024","EMESET 2ML INJ (2 MG/1 ML) ",13.35 , 4, '-',198.00,198.00],
        ["27/11/2024","MEROZA 500MG INJ ",809.77 , 2, '-',1619.54,1619.54],
        ["27/11/2024","METRONIDAZOLE 100ML IV (NIRLIFE) ",23.52 , 1, '-',23.52,23.52],
        ["27/11/2024","PARACETAMOL (BBRAUN)100ML IV IN) ",654.50 , 1, '-',654.50,654.50],
        ["27/11/2024","RLOC 2ML INJ (25MG/ML) ",7.26 , 2, '-',14.52,14.52],
        ["27/11/2024","AMIKAMAC 100MG/2ML INJ VIAL ",34.21 , 1, '-',34.21,34.21],
        ["28/11/2024","AMIKAMAC 100MG/2ML INJ VIAL ",34.21 , 1, '-',34.21,34.21],
        ["28/11/2024","MEROZA 500MG INJ ",809.77 , 2, '-',809.77,809.77],
        ["29/11/2024","AMIKAMAC 100MG/2ML INJ VIAL ",34.21 , 1, '-',34.21,34.21],
        ["30/11/2024","AMIKAMAC 100MG/2ML INJ VIAL ",34.21 , 1, '-',34.21,34.21],
        ["30/11/2024","SODIUM CHLORIDE 0.9% 100ML (NIRLIFE)",22.03 , 3, '-',66.09,66.09],
        ["30/11/2024","PARACETAMOL (BBRAUN)100ML IV IN) ",654.50 , 1, '-',654.50,654.50],
        ["30/11/2024","METRONIDAZOLE 100ML IV (NIRLIFE) ",23.52 , 1, '-',23.52,23.52],
        ["30/11/2024","MEROZA 500MG INJ ",809.77 , 1, '-',809.77,809.77],
        ["30/11/2024","RLOC 2ML INJ (25MG/ML) ",7.26 , 1, '-',7.26,7.26],
        ["30/11/2024","EMESET 2ML INJ (2 MG/1 ML)  ",13.35 , 4, '-',53.40,53.40],
        ["30/11/2024","LASIX 40MG/ 4ML INJ (10 MG/1 ML). ",13.39 , 1, '-',13.39,13.39],



        ['23/11/2024', '2 WAY FOLLY CATHETER NO 8 ROMSON ', 442.0, 1,'-', 442.0, 442.0],
        ['23/11/2024', '3 WAY CONNECTOR WITH 10 CM EXT. ROMSON ', 360.0, 1,'-', 360.0, 360.0],
        ['23/11/2024 ', 'COSMETIC POUCH ORDINARY ', 106.0, 1,'-', 'Not admisible', 0],
        ['23/11/2024 ', 'ECG ELECTRODE PAED/NEONATAL (ICONET) ', 42.0, 5,'-', 210.0, 210.0],
        ['23/11/2024 ', 'EXAMINATION GLOVES (NON STERILE) LATEX ', 18.0, 6,'-', 108.0, 108.0],
        ['23/11/2024 ', 'GAMJEE PAD 10INCH X 5INCH ', 40.0, 5,'-', 200.0, 200.0],
        ['23/11/2024 ', 'GAMJEE PAD 6INCH X 5INCH ', 30.0, 5,'-', 150.0, 150.0],
        ['23/11/2024 ', 'GAUZE SWAB (2 S) ', 10.0, 5,'-', 50.0, 50.0],
        ['23/11/2024 ', 'LANCET NEEDLE 30G ', 4.0, 3,'-', 12.0, 12.0],
        ['23/11/2024 ', 'NEOFLON 24G BD. ', 379.00, 2,'-', 758.0, 758.0],
        ['23/11/2024 ', 'OUNCE GLASS ', 50.0, 1,'-', 50.0, 50.0],
        ['23/11/2024 ', 'PM o LINE 200 CM MF ROMSON ', 425.0, 4,'-', 1700.0, 1700.0],
        ['23/11/2024 ', 'PAPER GLOVES ASPIRE ', 40.0, 5,'-', 200.0, 200.0],
        ['23/11/2024 ', 'RYLES TUBE BG ROMSON ', 77.0, 1,'-', 77.0, 77.0],
        ['23/11/2024 ', 'SYRINGES 1 ML DISPOVAN ', 10.0, 2,'-', 20.0, 20.0],
        ['23/11/2024 ', 'SYRINGES 10 ML DISPOVAN ', 13.0, 5,'-', 65.0, 65.0],
        ['23/11/2024 ', 'SYRINGES 2/2.5 ML DISPOVAN ', 5.40, 5,'-', 27.0, 27.0],
        ['23/11/2024 ', 'SYRINGES 50 ML DISPOVAN ', 61.0, 2,'-', 122.0, 122.0],
        ['23/11/2024 ', 'UNDER PAD 90X60 AQUAMED ', 135.00, 4,'-', 'Not admisible',0],
        ['23/11/2024 ', 'UROBAG (ROMSON) ', 360.0, 1,'-', 360.0, 360.0],
        ['23/11/2024 ', 'UROMETER ROMSON ', 572.0, 1,'-', 572.0, 572.0],
        ['23/11/2024 ', 'SYRINGES 5 ML DISPOVAN ', 10.0, 5,'-', 50.0, 50.0],
        ['23/11/2024 ', 'TEGADERM 6CM X 7CM VELFIX ', 181.0, 1,'-', 181.0, 181.0],
        ['23/11/2024 ', 'DIGITAL THERMOMETER MT100-DR MOREPEN ', 170.0, 1,'-', 170.0, 170.0],
        ['23/11/2024 ', 'ACCU-CHEK INSTANT 100 CT STRIP (ROCHE) ', 19.98, 3,'-', 59.94, 59.94],
        ['23/11/2024 ', 'PM 0 LINE 200 CM MF ROMSON ', 425.0, 3,'-', 1275.0, 1275.0],
        ['23/11/2024 ', 'SYRINGES 20 ML -BD ', 54.0, 3,'-', 162.0, 162.0],
        ['24/11/2024 ', '2 WAY FOLLY CATHETER NO 10 ROMSON ', 442.0, 1,'-', 442.0, 442.0],
        ['24/11/2024 ', 'SYRINGES 212.5 ML DISPOVAN ', 5.4, 5,'-', 27.0, 27.0],
        ['24/11/2024 ', 'UNDER PAD 90X60 AQUAMED ', 135.0, 4,'-', 'Not admisible', 0],
        ['24/11/2024 ', 'SYRINGES 20 ML .BD pal ', 44.0, 4,'-', 216.0, 216.0],
        ['24/11/2024 ', 'GAMJEE PAD 101NC1 X SINCH souvik ', 40.0, 4,'-', 160.0, 160.0],
        ['24/11/2024 ', 'GAMJEE PAD GINCH X SINCH ', 30.0, 4,'-', 120.0, 120.0],
        ['24/11/2024 ', 'GAUZE SWAB (2 5) ', 10.0, 5,'-', 50.0, 50.0],
        ['24/11/2024 ', 'PAPER GLOVES ASPIRE ', 40.0, 4,'-', 160.0, 160.0],
        ['24/11/2024 ', 'SYRINGES 1 ML DISPOVAN ', 10.0, 2,'-', 20.0, 20.0],
        ['24/11/2024 ', 'SYRINGES 2/2.5 ML DISPOVAN ', 5.4, 2,'-', 18.8, 18.8],
        ['24/11/2024 ', 'UNDER PAD 90X60 AQUAMED ', 135.0, 5,'-', 'Not admisible', 0.0],
        ['24/11/2024 ', 'SYRINGES 20 ML -BD ', 54.0, 3, 162.0,'-', 162.0],
        ['25/11/2024 ', 'RYLES TUBE 8G ROMSON ', 77.0, 1,'-', 77.0, 77.0],
        ['25/11/2024 ', '3 WAY CONNECTOR WITH 10 CM EXT. - ROMSON ', 360.0, 1,'-', 360.0, 360.0],
        ['25/11/2024 ', 'ALCO SWAB ROMSON ', 2.75, 1,'-', 2.25, 2.25],
        ['25/11/2024 ', 'ECG ELECTRODE PAED/NEONATAL (ICONET) ', 42.0, 3,'-', 126.0, 126.0],
        ['25/11/2024 ', 'GAMJEE PAD 10INCH SINCH ', 40.0, 4,'-', 160.0, 160.0],
        ['25/11/2024 ', 'GAMJEE PAD 6INCH X 51NCH ', 30.0, 4,'-', 120.0, 120.0],
        ['25/11/2024 ', 'NEOFLON 24G BD. ', 379.0, 1,'-', 379.0, 379.0],
        ['25/11/2024 ', 'SYRINGES 10 ML DISPOVAN ', 13.0, 3,'-', 39.0, 39.0],
        ['25/11/2024 ', 'SYRINGES 2/2.5 ML . DISPOVAN ', 5.4, 3,'-', 15.2, 15.2],
        ['25/11/2024 ', 'SYRINGES 50 ML DISPOVAN ', 61.0, 3,'-', 193.0, 193.0],
        ['25/11/2024 ', 'UNDER PAD 90X60 AQUAMED ', 135.00, 4,'-', 'Not admisible', 0.0],
        ['25/11/2024 ', 'SYRINGES 5 ML . DISPOVAN ', 10.00, 3,'-', 30.0, 30.0],
        ['25/11/2024 ', 'TEGADERM 7CM X 9CM VELFIX (1633) ', 189.0, 1,'-', 189.0, 189.0],
        ['26/11/2024 ', 'GAMJEE PAD 10JNCH X 51NCH ', 40.0, 3,'-', 120.0, 120.0],
        ['26/11/2024 ', 'GAMJEE PAD 6INCH X SINCH ', 30.0, 3,'-', 90.0, 90.0],
        ['26/11/2024 ', 'SYRINGES 2/2.5 ML - DISPOVAN ', 5.4, 2,'-', 10.8, 10.8],
        ['26/11/2024 ', 'UNDER PAD 90X60 - AQUAMED ', 135.0, 1,'-', 0.0, 0.0],
        ['26/11/2024 ', 'SYRINGES 20 ML -BD ', 54.0, 2,'-', 108.0, 108.0],
        ['26/11/2024 ', 'COTTON 400 GM GROSS ', 310.0, 1,'-', 310.0, 310.0],
        ['26/11/2024 ', 'SYRINGES 20 ML-BD ', 54.0, 5,'-', 270.0, 270.0],
        ['27/11/2024 ', 'GAMJEE PAD 10INCH X SINCH ', 40.0, 1,'-', 40.0, 40.0],
        ['27/11/2024 ', 'GAMJEE PAD 6INCH X 5INCH ', 30.0, 4,'-', 120.0, 120.0],
        ['27/11/2024 ', 'SYRINGES 2/2.5 ML DISPOVAN ', 5.4, 4,'-', 21.6, 21.6],
        ['27/11/2024 ', 'UNDER PAD 90X60 AQUAMED ', 135.0, 5,'-', 'Not admisible', 0.0],
        ['27/11/2024 ', 'SYRINGES 5 ML - DISPOVAN ', 10.0, 1,'-', 10.0, 10.0],
        ['27/11/2024 ', 'SYRINGES 20 ML -BD ', 54.0, 5,'-', 270.0, 270.0],
        ['28/11/2024 ', 'GAMJEE PAD 6INCH X 51NCH ', 30.0, 4,'-', 120.0, 120.0],
        ['28/11/2024 ', 'SYRINGES 2/2.5 ML - DISPOVAN ', 5.4, 2,'-', 10.8, 10.8],
        ['28/11/2024 ', 'UNDER PAD 90X60 - AQUAMED ', 135.0, 5,'-', "'Not admisible'", 0.0],
        ['28/11/2024 ', 'SYRINGES 20 ML-BD ', 54.0, 2,'-', 108.0, 108.0],
        ['29/11/2024 ', 'UNDER PAD 90X60 AQUAMED ', 135.0, 5,'-', 'Not admisible', 0.0],
        ['29/11/2024 ', 'SYRINGES 20 ML -BD ', 54.0, 3,'-', 162.0, 162.0],
        ['29/11/2024 ', 'ALCO SWAB ROMSON ', 2.75, 1,'-', 2.75, 2.75],
        ['29/11/2024 ', 'NEOFLON 24G 8D. ', 379.0, 1,'-', 379.0, 379.0],
        ['29/11/2024 ', 'NEOFLON 26G ', 379.0, 1,'-', 379.0, 379.0],
        ['29/11/2024 ', 'TEGADERM 6CM X 7CM VELFIX ', 181.0, 1,'-', 181.0, 181.0],
        ['30/11/2024 ', 'GAFDEE PAD 101NCM X SINCE ', 40.0, 2,'-', 80.0, 80.0],
        ['30/11/2024', 'SYRINGES 20 ML .BD ', 54.0, 2,'-', 108.0, 108.0],
        ['30/11/2024 ', 'UNDER PAD 90x60 AQUAMED ', 135.0, 1,'-', 'Not admisible', 0.0],
        ['30/11/2024', 'PM.0 LINE 200 CM MF ROMSON ', 425.0, 4,'-', 1700.0, 1700.0],
        ['01/11/2014 ', 'STRINGES 10 Mr. , DISPOVAN ', 13.0, 2,'-', 26.0, 26.0],
        ['01/11/2024 ', 'SYRINGES 2/2.5 ML DISPOVAN ', 5.4, 4,'-', 20.16, 20.16],
        ['01/12/2024', 'GAMJEE PAD 6INCH X 5INCH ', 30.0, 4,'-', 120.0, 120.0],
        ['01/12/2024 ', 'GAMJEE PAD 6INCH X 5INCH ', 30.0, 4,'-', 120.0, 120.0],
        ['01/12/2014 ', 'GAMJEE PAD 6INCH X 5INCH ', 30.0, 4,'-', 120.0, 120.0],
        ['01/12/2004 ', 'GAFDEE PAD 101NCM X SINCE ', 40.0, 4,'-', 160.0, 160.0],



        ["23/11/2024","Ryles tube insertion", 550,1, 500, 500, 500],
        ["24/11/2024","Ryles tube insertion", 550,1, 500, 500, 500],

        ["23/11/2024","Foley`s catherter insertion", 550,1, 700, 500, 500],
        ["24/11/2024","Foley`s catherter insertion", 550,1, 700, 500, 500],

        ["26/11/2024","Syring Pump", 550,1, 450, 450, 450],
        ["27/11/2024","Syring Pump", 550,1, 450, 450, 450],
        ["28/11/2024","Syring Pump", 550,1, 450, 450, 450],
        ["29/11/2024","Syring Pump", 550,1, 450, 450, 450],
        ["30/11/2024","Syring Pump", 550,1, 450, 450, 450],
        ["01/12/2024","Syring Pump", 550,1, 450, 450, 450],


        ["23/11/2024","MRD charge", 1000,1, 300, 300, 300],

        ["23/11/2024","VBG", 1210,1, '-', 1210, 1210],
        ["23/11/2024","CBS", 220,1, '-', 220, 220],
        ["24/11/2024","CBS", 220,1, '-', 220, 220],
        ["25/11/2024","CBS", 220,1, '-', 220, 220],
        ["26/11/2024","CBS", 220,1, '-', 220, 220]
    ]
    
    specified_row = 8
    
    # Write data to Excel at the specified row
    for row_index, row_data in enumerate(data, start=specified_row):
        for col_index, value in enumerate(row_data, start=3):  # Columns start from 1 (A)
            ws.cell(row=row_index, column=col_index, value=value)
        
    ws['D200']='Total amount Claimed'
    ws['E200']=110017
    
    
    ws['G200']='Total amount Admisble'
    ws['H200']=100826
    
    
    
    
 
 
    wb.save(r"G:\DFCCIL\Sheet\final.xlsx")
